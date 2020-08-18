from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import argparse

parser = argparse.ArgumentParser(description="주민번호 탐색기")
parser.add_argument('--file', required=True, help='file name')
parser.add_argument('--sheet', required=True, type=int,  help='sheet 위치번호')
parser.add_argument('--col', required=True, help='검색할 Column 위치 ex) A B C')

args = parser.parse_args()
global c
c=0

def reg_num(cel_data, num):
    global c
    tmp = cel_data.split(" ")
    reg_result = 0
    for data in tmp:
        if data.isdigit() and len(data)==13:
            birth_month = data[2:4]
            birth_day = data[4:6]
            sex = data[6]
            if int(birth_day) < 32 and int(birth_month) < 13 and 5 > int(sex) > 0:
                reg_result = 1
                print("라인 : " + str(num) + " 검출 주민번호 : " + data)
                c += 1
    return reg_result


def search_columnlist(filename):
    #wb = openpyxl.load_workbook("SAMPLE.xlsx", read_only=False, keep_vba=True)
    wb = load_workbook(filename)
    # print(wb.sheetnames[1])
    sel_sheet = wb.sheetnames[int(args.sheet)-1]
    wb.active
    sheet = wb[sel_sheet]
    search_column = sheet[args.col]
    search_cnt = 0
    font_c = Font(color="FF0000")
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    for i in search_column:
        search_cnt += 1
        if i.value:
            ch = reg_num(i.value, search_cnt)
            if ch:
                check_num_color = sheet.cell(row=search_cnt, column=3)
                check_num_color.fill = redFill

    wb.save("result.xlsx")


def main():
    global c
    c = 0 #탐지 총 건수
    search_columnlist(args.file)
    print("총 검출 수 : " + str(c))


if __name__ == "__main__":
    main()
